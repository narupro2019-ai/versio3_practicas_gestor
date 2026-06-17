from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
import psycopg2
from psycopg2.extras import RealDictCursor
from werkzeug.security import generate_password_hash, check_password_hash
import os
import pandas as pd
import io
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'practicas-secret-2026')

DATABASE_URL = os.environ.get('DATABASE_URL')

def get_db_connection():
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL no está configurada en las variables de entorno")
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            nombre TEXT NOT NULL,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS estudiantes (
            id SERIAL PRIMARY KEY,
            cedula TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            sitio TEXT,
            programa TEXT DEFAULT 'Fisioterapia',
            sede TEXT,
            nivel_practica TEXT,
            grupo TEXT,
            correo TEXT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS docentes (
            id SERIAL PRIMARY KEY,
            documento TEXT UNIQUE,
            nombre TEXT NOT NULL,
            correo TEXT,
            estado TEXT DEFAULT 'Activo'
        );

        CREATE TABLE IF NOT EXISTS escenarios (
            id SERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            direccion TEXT,
            cupos INTEGER DEFAULT 10,
            estado TEXT DEFAULT 'Activo'
        );

        CREATE TABLE IF NOT EXISTS asignaciones (
            id SERIAL PRIMARY KEY,
            estudiante_id INTEGER REFERENCES estudiantes(id) ON DELETE CASCADE,
            docente_id INTEGER REFERENCES docentes(id) ON DELETE SET NULL,
            escenario_id INTEGER REFERENCES escenarios(id) ON DELETE SET NULL,
            nivel_practica TEXT,
            grupo TEXT,
            rotacion INTEGER NOT NULL,
            horario TEXT,
            fecha_inicio DATE,
            fecha_fin DATE,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    conn.commit()

    # Migración segura: agrega la columna 'predeterminado' si la tabla usuarios
    # ya existía de una versión anterior sin esta columna
    cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS predeterminado BOOLEAN DEFAULT FALSE")
    conn.commit()

    # Crear usuario administrador por defecto si todavía no existe ningún usuario
    cur.execute("SELECT COUNT(*) AS total FROM usuarios")
    if cur.fetchone()['total'] == 0:
        admin_user = os.environ.get('ADMIN_USERNAME', 'admin')
        admin_pass = os.environ.get('ADMIN_PASSWORD', 'admin123')
        cur.execute('''
            INSERT INTO usuarios (username, password_hash, nombre, predeterminado)
            VALUES (%s, %s, %s, TRUE)
        ''', (admin_user, generate_password_hash(admin_pass), 'Administrador'))
        conn.commit()

    cur.close()
    conn.close()

with app.app_context():
    init_db()

# ==================== AUTENTICACIÓN ====================

# Endpoints accesibles sin haber iniciado sesión
PUBLIC_ENDPOINTS = {'login', 'static'}

@app.before_request
def require_login():
    if request.endpoint in PUBLIC_ENDPOINTS or request.endpoint is None:
        return
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios WHERE username = %s", (username,))
        usuario = cur.fetchone()
        cur.close()
        conn.close()

        if usuario and check_password_hash(usuario['password_hash'], password):
            session['user_id'] = usuario['id']
            session['username'] = usuario['username']
            session['nombre'] = usuario['nombre']
            flash(f"✅ Bienvenido, {usuario['nombre']}", 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('⚠️ Usuario o contraseña incorrectos', 'danger')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('👋 Sesión cerrada correctamente', 'success')
    return redirect(url_for('login'))

# ==================== USUARIOS CRUD ====================
@app.route('/usuarios')
def usuarios():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM usuarios ORDER BY nombre")
    usuarios = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/register_usuario', methods=['GET', 'POST'])
def register_usuario():
    if request.method == 'POST':
        username = request.form['username'].strip()
        nombre = request.form['nombre'].strip()
        password = request.form['password']

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute('''
                INSERT INTO usuarios (username, password_hash, nombre)
                VALUES (%s, %s, %s)
            ''', (username, generate_password_hash(password), nombre))
            conn.commit()

            # Si existe el usuario predeterminado creado automáticamente al
            # iniciar la app, se elimina al crear el primer usuario real
            cur.execute(
                "SELECT id FROM usuarios WHERE predeterminado = TRUE AND username != %s",
                (username,)
            )
            predeterminado = cur.fetchone()
            if predeterminado:
                cur.execute("DELETE FROM usuarios WHERE id = %s", (predeterminado['id'],))
                conn.commit()
                flash('✅ Usuario creado con éxito. El usuario "admin" predeterminado fue eliminado por seguridad', 'success')
            else:
                flash('✅ Usuario creado con éxito', 'success')

            return redirect(url_for('usuarios'))
        except psycopg2.IntegrityError:
            flash('⚠️ Ya existe un usuario con ese nombre de usuario', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    return render_template('register_usuario.html')

@app.route('/edit_usuario/<int:id>', methods=['GET', 'POST'])
def edit_usuario(id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        username = request.form['username'].strip()
        nombre = request.form['nombre'].strip()
        password = request.form.get('password', '').strip()

        try:
            if password:
                cur.execute('''
                    UPDATE usuarios SET username=%s, nombre=%s, password_hash=%s WHERE id=%s
                ''', (username, nombre, generate_password_hash(password), id))
            else:
                cur.execute('''
                    UPDATE usuarios SET username=%s, nombre=%s WHERE id=%s
                ''', (username, nombre, id))
            conn.commit()

            # Si el usuario editado es el que tiene la sesión activa, refrescar sus datos en sesión
            if session.get('user_id') == id:
                session['username'] = username
                session['nombre'] = nombre

            flash('✅ Usuario actualizado', 'success')
            return redirect(url_for('usuarios'))
        except psycopg2.IntegrityError:
            flash('⚠️ Ya existe un usuario con ese nombre de usuario', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()
        return redirect(url_for('edit_usuario', id=id))

    cur.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
    usuario = cur.fetchone()
    cur.close()
    conn.close()

    if not usuario:
        flash('Usuario no encontrado', 'danger')
        return redirect(url_for('usuarios'))

    return render_template('edit_usuario.html', usuario=usuario)

@app.route('/delete_usuario/<int:id>')
def delete_usuario(id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) AS total FROM usuarios")
    total = cur.fetchone()['total']

    if total <= 1:
        flash('⚠️ No se puede eliminar el único usuario del sistema', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('usuarios'))

    cur.execute("DELETE FROM usuarios WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()

    flash('🗑️ Usuario eliminado', 'danger')

    if session.get('user_id') == id:
        session.clear()
        flash('Tu sesión se cerró porque eliminaste tu propio usuario', 'warning')
        return redirect(url_for('login'))

    return redirect(url_for('usuarios'))

# ==================== DASHBOARD ====================
@app.route('/')
def index():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT a.id, e.nombre as estudiante, e.cedula, d.nombre as docente, 
               es.nombre as escenario, a.rotacion, a.horario, a.fecha_inicio, a.fecha_fin
        FROM asignaciones a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN docentes d ON a.docente_id = d.id
        JOIN escenarios es ON a.escenario_id = es.id
        ORDER BY a.fecha_creacion DESC LIMIT 10
    ''')
    asignaciones = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('index.html', asignaciones=asignaciones)

# ==================== ESTUDIANTES CRUD ====================
@app.route('/estudiantes')
def estudiantes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM estudiantes ORDER BY nombre")
    estudiantes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('estudiantes.html', estudiantes=estudiantes)

@app.route('/register_estudiante', methods=['GET', 'POST'])
def register_estudiante():
    if request.method == 'POST':
        cedula = request.form['cedula'].strip()
        nombre = request.form['nombre'].strip()
        sitio = request.form['sitio'].strip()
        nivel_practica = request.form['nivel_practica']
        programa = request.form.get('programa', 'Fisioterapia')
        sede = request.form['sede'].strip()
        correo = request.form.get('correo', '').strip()

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute('''
                INSERT INTO estudiantes (cedula, nombre, sitio, nivel_practica, programa, sede, correo)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (cedula, nombre, sitio, nivel_practica, programa, sede, correo))
            conn.commit()
            flash('✅ Estudiante registrado con éxito', 'success')
            return redirect(url_for('estudiantes'))
        except psycopg2.IntegrityError:
            flash('⚠️ Ya existe un estudiante con esa cédula', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    return render_template('register_estudiante.html')

@app.route('/edit_estudiante/<int:id>', methods=['GET', 'POST'])
def edit_estudiante(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        cedula = request.form['cedula'].strip()
        nombre = request.form['nombre'].strip()
        sitio = request.form['sitio'].strip()
        nivel_practica = request.form['nivel_practica']
        programa = request.form.get('programa', 'Fisioterapia')
        sede = request.form['sede'].strip()
        correo = request.form.get('correo', '').strip()

        cur.execute('''
            UPDATE estudiantes 
            SET cedula=%s, nombre=%s, sitio=%s, nivel_practica=%s, 
                programa=%s, sede=%s, correo=%s
            WHERE id=%s
        ''', (cedula, nombre, sitio, nivel_practica, programa, sede, correo, id))
        conn.commit()
        flash('✅ Estudiante actualizado', 'success')
        return redirect(url_for('estudiantes'))

    cur.execute("SELECT * FROM estudiantes WHERE id = %s", (id,))
    estudiante = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('edit_estudiante.html', estudiante=estudiante)

@app.route('/delete_estudiante/<int:id>')
def delete_estudiante(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM estudiantes WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('🗑️ Estudiante eliminado', 'danger')
    return redirect(url_for('estudiantes'))

# ==================== DOCENTES CRUD ====================
@app.route('/docentes')
def docentes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM docentes ORDER BY nombre")
    docentes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('docentes.html', docentes=docentes)

@app.route('/register_docente', methods=['GET', 'POST'])
def register_docente():
    if request.method == 'POST':
        documento = request.form['documento'].strip()
        nombre = request.form['nombre'].strip()
        correo = request.form.get('correo', '').strip()

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute('''
                INSERT INTO docentes (documento, nombre, correo)
                VALUES (%s, %s, %s)
            ''', (documento, nombre, correo))
            conn.commit()
            flash('✅ Docente registrado con éxito', 'success')
            return redirect(url_for('docentes'))
        except psycopg2.IntegrityError:
            flash('⚠️ Ya existe un docente con ese documento', 'danger')
        finally:
            cur.close()
            conn.close()
    return render_template('register_docente.html')

@app.route('/edit_docente/<int:id>', methods=['GET', 'POST'])
def edit_docente(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        documento = request.form['documento'].strip()
        nombre = request.form['nombre'].strip()
        correo = request.form.get('correo', '').strip()

        cur.execute('''
            UPDATE docentes SET documento=%s, nombre=%s, correo=%s WHERE id=%s
        ''', (documento, nombre, correo, id))
        conn.commit()
        flash('✅ Docente actualizado', 'success')
        return redirect(url_for('docentes'))

    cur.execute("SELECT * FROM docentes WHERE id = %s", (id,))
    docente = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('edit_docente.html', docente=docente)

@app.route('/delete_docente/<int:id>')
def delete_docente(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM docentes WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('🗑️ Docente eliminado', 'danger')
    return redirect(url_for('docentes'))

# ==================== ESCENARIOS CRUD ====================
@app.route('/escenarios')
def escenarios():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM escenarios ORDER BY nombre")
    escenarios = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('escenarios.html', escenarios=escenarios)

@app.route('/register_escenario', methods=['GET', 'POST'])
def register_escenario():
    if request.method == 'POST':
        nombre = request.form['nombre'].strip()
        direccion = request.form.get('direccion', '').strip()
        cupos = int(request.form.get('cupos', 10))

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO escenarios (nombre, direccion, cupos)
            VALUES (%s, %s, %s)
        ''', (nombre, direccion, cupos))
        conn.commit()
        cur.close()
        conn.close()
        flash('✅ Escenario registrado', 'success')
        return redirect(url_for('escenarios'))
    return render_template('register_escenario.html')

@app.route('/edit_escenario/<int:id>', methods=['GET', 'POST'])
def edit_escenario(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        nombre = request.form['nombre'].strip()
        direccion = request.form.get('direccion', '').strip()
        cupos = int(request.form.get('cupos', 10))

        cur.execute('''
            UPDATE escenarios SET nombre=%s, direccion=%s, cupos=%s WHERE id=%s
        ''', (nombre, direccion, cupos, id))
        conn.commit()
        flash('✅ Escenario actualizado', 'success')
        return redirect(url_for('escenarios'))

    cur.execute("SELECT * FROM escenarios WHERE id = %s", (id,))
    escenario = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('edit_escenario.html', escenario=escenario)

@app.route('/delete_escenario/<int:id>')
def delete_escenario(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM escenarios WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('🗑️ Escenario eliminado', 'danger')
    return redirect(url_for('escenarios'))

# ==================== ASIGNACIONES - CRUD COMPLETO ====================

@app.route('/asignaciones')
def asignaciones_list():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT 
            a.id, 
            e.nombre as estudiante, 
            e.cedula,
            d.nombre as docente, 
            es.nombre as escenario, 
            a.rotacion, 
            a.horario, 
            a.fecha_inicio, 
            a.fecha_fin
        FROM asignaciones a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN docentes d ON a.docente_id = d.id
        JOIN escenarios es ON a.escenario_id = es.id
        ORDER BY a.fecha_creacion DESC
    ''')
    asignaciones = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('asignaciones.html', asignaciones=asignaciones)


@app.route('/new_assignment', methods=['GET', 'POST'])
def new_assignment():
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        try:
            estudiante_id = int(request.form['estudiante_id'])
            docente_id = int(request.form['docente_id'])
            escenario_id = int(request.form['escenario_id'])
            rotacion = int(request.form['rotacion'])
            horario = request.form['horario'].strip()
            fecha_inicio = request.form['fecha_inicio']
            fecha_fin = request.form['fecha_fin']

            # Validación de conflictos (incluyendo horario)
            cur.execute('''
                SELECT COUNT(*) AS count FROM asignaciones 
                WHERE estudiante_id = %s 
                  AND horario = %s 
                  AND ((fecha_inicio <= %s AND fecha_fin >= %s) 
                    OR (fecha_inicio <= %s AND fecha_fin >= %s))
            ''', (estudiante_id, horario, fecha_fin, fecha_inicio, fecha_inicio, fecha_fin))
            
            if cur.fetchone()['count'] > 0:
                flash('❌ Conflicto detectado: El estudiante ya tiene asignación en ese horario y fechas', 'danger')
                cur.close()
                conn.close()
                return redirect(url_for('new_assignment'))

            # Inserción en la tabla
            cur.execute('''
                INSERT INTO asignaciones (estudiante_id, docente_id, escenario_id, rotacion, horario, fecha_inicio, fecha_fin)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (estudiante_id, docente_id, escenario_id, rotacion, horario, fecha_inicio, fecha_fin))
            conn.commit()
            flash('✅ Asignación creada correctamente', 'success')
            return redirect(url_for('asignaciones_list'))

        except Exception as e:
            flash(f'Error al guardar: {str(e)}', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()

    # GET - cargar listas
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, cedula FROM estudiantes ORDER BY nombre")
    estudiantes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM docentes WHERE estado = 'Activo' ORDER BY nombre")
    docentes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM escenarios WHERE estado = 'Activo' ORDER BY nombre")
    escenarios = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template('new_assignment.html', estudiantes=estudiantes, docentes=docentes, escenarios=escenarios)


@app.route('/edit_assignment/<int:id>', methods=['GET', 'POST'])
def edit_assignment(id):
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        try:
            estudiante_id = int(request.form['estudiante_id'])
            docente_id = int(request.form['docente_id'])
            escenario_id = int(request.form['escenario_id'])
            rotacion = int(request.form['rotacion'])
            horario = request.form.get('horario', '').strip()
            fecha_inicio = request.form['fecha_inicio']
            fecha_fin = request.form['fecha_fin']

            # Validación de conflictos (incluyendo horario)
            cur.execute('''
                SELECT COUNT(*) AS count FROM asignaciones 
                WHERE estudiante_id = %s 
                  AND horario = %s 
                  AND id <> %s
                  AND ((fecha_inicio <= %s AND fecha_fin >= %s) 
                    OR (fecha_inicio <= %s AND fecha_fin >= %s))
            ''', (estudiante_id, horario, id, fecha_fin, fecha_inicio, fecha_inicio, fecha_fin))
            
            if cur.fetchone()['count'] > 0:
                flash('❌ Conflicto detectado: El estudiante ya tiene asignación en ese horario y fechas', 'danger')
                cur.close()
                conn.close()
                return redirect(url_for('edit_assignment', id=id))

            # ✅ Actualización de la asignación seleccionada
            cur.execute('''
                UPDATE asignaciones 
                SET estudiante_id = %s,
                    docente_id = %s,
                    escenario_id = %s,
                    rotacion = %s,
                    horario = %s,
                    fecha_inicio = %s,
                    fecha_fin = %s
                WHERE id = %s
            ''', (estudiante_id, docente_id, escenario_id, rotacion, 
                  horario, fecha_inicio, fecha_fin, id))

            # ✅ Sincronizar fechas y horario en toda la rotación
            cur.execute('''
                UPDATE asignaciones
                SET fecha_inicio = %s,
                    fecha_fin = %s,
                    horario = %s
                WHERE rotacion = %s
            ''', (fecha_inicio, fecha_fin, horario, rotacion))

            conn.commit()
            flash(f'✅ Asignación actualizada y fechas/horario sincronizados en Rotación {rotacion}', 'success')
            return redirect(url_for('asignaciones_list'))
            
        except Exception as e:
            flash(f'❌ Error al actualizar: {str(e)}', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()

    # ==================== GET - Cargar datos para editar ====================
    cur = conn.cursor()
    cur.execute('SELECT * FROM asignaciones WHERE id = %s', (id,))
    asignacion = cur.fetchone()

    if not asignacion:
        flash('Asignación no encontrada', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('asignaciones_list'))

    # Cargar listas para los selects
    cur.execute("SELECT id, nombre, cedula FROM estudiantes ORDER BY nombre")
    estudiantes = cur.fetchall()
    
    cur.execute("SELECT id, nombre FROM docentes WHERE estado = 'Activo' ORDER BY nombre")
    docentes = cur.fetchall()
    
    cur.execute("SELECT id, nombre FROM escenarios WHERE estado = 'Activo' ORDER BY nombre")
    escenarios = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template('edit_assignment.html', 
                         asignacion=asignacion, 
                         estudiantes=estudiantes, 
                         docentes=docentes, 
                         escenarios=escenarios)




@app.route('/delete_assignment/<int:id>')
def delete_assignment(id):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Verificar que la asignación existe antes de eliminar
        cur.execute("SELECT id FROM asignaciones WHERE id = %s", (id,))
        if not cur.fetchone():
            flash('Asignación no encontrada', 'warning')
            return redirect(url_for('asignaciones_list'))

        cur.execute("DELETE FROM asignaciones WHERE id = %s", (id,))
        conn.commit()
        
        flash('🗑️ Asignación eliminada correctamente', 'danger')
        return redirect(url_for('asignaciones_list'))
        
    except Exception as e:
        flash(f'❌ Error al eliminar la asignación: {str(e)}', 'danger')
        if conn:
            conn.rollback()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

# ==================== REPORTES ====================

# ==================== EXPORTAR A EXCEL ====================
@app.route('/generate_excel_report')
def generate_excel_report():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT 
            e.nombre AS "Estudiante",
            e.cedula AS "Documento",
            es.nombre AS "Escenario",
            d.nombre AS "Docente",
            a.rotacion AS "Rotación",
            a.horario AS "Horario",
            a.fecha_inicio AS "Fecha Inicio",
            a.fecha_fin AS "Fecha Fin",
            es.direccion AS "Dirección"
        FROM asignaciones a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN docentes d ON a.docente_id = d.id
        JOIN escenarios es ON a.escenario_id = es.id
        ORDER BY e.nombre ASC, a.rotacion ASC
    ''')
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        flash('⚠️ No hay asignaciones para exportar', 'warning')
        return redirect(url_for('index'))

    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    columns = ["Estudiante", "Documento", "Escenario", "Docente", "Rotación",
               "Horario", "Fecha Inicio", "Fecha Fin", "Dirección"]
    df = pd.DataFrame(rows, columns=columns)
    for col in ["Fecha Inicio", "Fecha Fin"]:
        df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime("%d/%m/%Y").fillna("")

    # ── Colores (mismos que el PDF) ──
    C_TITULO    = "2E75B6"
    C_ROTACION  = "1F4E79"
    C_ESCENARIO = "003366"
    C_DOCENTE   = "E2EFDA"
    C_GRIS      = "F2F2F2"
    C_BLANCO    = "FFFFFF"

    thin   = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font_white_bold(size=10):
        return Font(name='Arial', bold=True, color='FFFFFF', size=size)

    def font_dark_bold_italic(size=9):
        return Font(name='Arial', bold=True, italic=True, color='1F4E79', size=size)

    def font_normal(size=9):
        return Font(name='Arial', size=size)

    def align_center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def align_left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Programación 2026-1"
    current_row = 1

    # ── Título principal ──
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    cell = ws.cell(row=current_row, column=1,
                   value="PROGRAMACIÓN DE PRÁCTICAS ACADÉMICAS 2026-1")
    cell.fill = fill(C_TITULO)
    cell.font = font_white_bold(13)
    cell.alignment = align_center()
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    for horario, df_horario in df.groupby("Horario", sort=True):

        # ── Banda de horario ──
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=8)
        cell = ws.cell(row=current_row, column=1, value=f"Horario: {horario}")
        cell.fill = fill(C_TITULO)
        cell.font = font_white_bold(11)
        cell.alignment = align_left()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for rotacion, df_rot in df_horario.groupby("Rotación", sort=True):
            fecha_ini = df_rot["Fecha Inicio"].iloc[0]
            fecha_fin = df_rot["Fecha Fin"].iloc[0]
            escenarios_orden = list(dict.fromkeys(df_rot["Escenario"].tolist()))
            n_cols = len(escenarios_orden)

            esc_data = {}
            for esc in escenarios_orden:
                sub = df_rot[df_rot["Escenario"] == esc]
                esc_data[esc] = {
                    "docente":     sub["Docente"].iloc[0],
                    "estudiantes": sub["Estudiante"].tolist()
                }

            # ── Banda de rotación ──
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=n_cols)
            cell = ws.cell(row=current_row, column=1,
                           value=f"Rotación {rotacion}:   {fecha_ini}  →  {fecha_fin}")
            cell.fill = fill(C_ROTACION)
            cell.font = font_white_bold(10)
            cell.alignment = align_left()
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            # ── Fila de escenarios ──
            for col_idx, esc in enumerate(escenarios_orden, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=esc)
                cell.fill = fill(C_ESCENARIO)
                cell.font = font_white_bold(9)
                cell.alignment = align_center()
                cell.border = border
            ws.row_dimensions[current_row].height = 30
            current_row += 1

            # ── Fila de docentes ──
            for col_idx, esc in enumerate(escenarios_orden, start=1):
                cell = ws.cell(row=current_row, column=col_idx,
                               value=f"Docente: {esc_data[esc]['docente']}")
                cell.fill = fill(C_DOCENTE)
                cell.font = font_dark_bold_italic(8)
                cell.alignment = align_center()
                cell.border = border
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            # ── Filas de estudiantes ──
            max_est = max(len(esc_data[e]["estudiantes"]) for e in escenarios_orden)
            for i in range(max_est):
                bg = C_GRIS if i % 2 == 0 else C_BLANCO
                for col_idx, esc in enumerate(escenarios_orden, start=1):
                    lst = esc_data[esc]["estudiantes"]
                    val = lst[i] if i < len(lst) else ""
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.fill = fill(bg)
                    cell.font = font_normal(8)
                    cell.alignment = align_center()
                    cell.border = border
                ws.row_dimensions[current_row].height = 14
                current_row += 1

            current_row += 1  # espacio entre rotaciones

    # ── Ancho de columnas ──
    for col_idx in range(1, 9):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Programacion_Practicas_2026-1.xlsx'
    )



@app.route('/generate_pdf_report')
def generate_pdf_report():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import pandas as pd
        import io

        # ✅ MISMA LÓGICA QUE EL EXCEL
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''
            SELECT 
                e.nombre AS "Estudiante",
                e.cedula AS "Documento",
                es.nombre AS "Escenario",
                d.nombre AS "Docente",
                a.rotacion AS "Rotación",
                a.horario AS "Horario",
                a.fecha_inicio AS "Fecha Inicio",
                a.fecha_fin AS "Fecha Fin",
                es.direccion AS "Dirección"
            FROM asignaciones a
            JOIN estudiantes e ON a.estudiante_id = e.id
            JOIN docentes d ON a.docente_id = d.id
            JOIN escenarios es ON a.escenario_id = es.id
            ORDER BY e.nombre ASC, a.rotacion ASC
        ''')
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if not rows:
            flash('⚠️ No hay asignaciones para exportar', 'warning')
            return redirect(url_for('index'))

        # ✅ MISMA LÓGICA QUE EL EXCEL
        columns = ["Estudiante", "Documento", "Escenario", "Docente", "Rotación",
                   "Horario", "Fecha Inicio", "Fecha Fin", "Dirección"]
        df = pd.DataFrame(rows, columns=columns)
        for col in ["Fecha Inicio", "Fecha Fin"]:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime("%d/%m/%Y").fillna("")

        # ── Colores ──
        AZUL_HEADER   = colors.HexColor('#003366')
        AZUL_ROTACION = colors.HexColor('#1F4E79')
        VERDE_DOCENTE = colors.HexColor('#E2EFDA')
        GRIS_ALTERNO  = colors.HexColor('#F2F2F2')
        AZUL_TITULO   = colors.HexColor('#2E75B6')
        BLANCO        = colors.white

        PAGE_W = landscape(letter)[0] - 2*cm
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                leftMargin=1*cm, rightMargin=1*cm,
                                topMargin=1.2*cm, bottomMargin=1.2*cm)
        styles = getSampleStyleSheet()
        elements = []

        # ── Estilos ──
        title_s = ParagraphStyle('t', parent=styles['Title'], fontName='Helvetica-Bold',
                                 fontSize=14, textColor=AZUL_TITULO, alignment=TA_CENTER, spaceAfter=4)
        rot_s   = ParagraphStyle('r', parent=styles['Normal'], fontName='Helvetica-Bold',
                                 fontSize=9, textColor=BLANCO, alignment=TA_LEFT)
        esc_s   = ParagraphStyle('e', parent=styles['Normal'], fontName='Helvetica-Bold',
                                 fontSize=8, textColor=BLANCO, alignment=TA_CENTER)
        doc_s   = ParagraphStyle('d', parent=styles['Normal'], fontName='Helvetica-BoldOblique',
                                 fontSize=7.5, textColor=colors.HexColor('#1F4E79'), alignment=TA_CENTER)
        est_s   = ParagraphStyle('s', parent=styles['Normal'], fontName='Helvetica',
                                 fontSize=7.5, alignment=TA_CENTER)

        elements.append(Paragraph("PROGRAMACIÓN DE PRÁCTICAS ACADÉMICAS 2026-1", title_s))
        elements.append(Spacer(1, 10))

        # ── Agrupar por Horario (grupo AM/PM) y luego por Rotación ──
        for horario, df_horario in df.groupby("Horario", sort=True):

            # Banda de grupo (AM / PM)
            grupo_tbl = Table([[Paragraph(f"Horario: {horario}", rot_s)]], colWidths=[PAGE_W])
            grupo_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,-1), AZUL_TITULO),
                ('TOPPADDING',    (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ]))
            elements.append(grupo_tbl)
            elements.append(Spacer(1, 6))

            for rotacion, df_rot in df_horario.groupby("Rotación", sort=True):
                fecha_ini = df_rot["Fecha Inicio"].iloc[0]
                fecha_fin = df_rot["Fecha Fin"].iloc[0]

                # Escenarios en el orden en que aparecen
                escenarios_orden = list(dict.fromkeys(df_rot["Escenario"].tolist()))

                esc_data = {}
                for esc in escenarios_orden:
                    sub = df_rot[df_rot["Escenario"] == esc]
                    esc_data[esc] = {
                        "docente":     sub["Docente"].iloc[0],
                        "estudiantes": sub["Estudiante"].tolist()
                    }

                n_cols = len(escenarios_orden)
                col_w  = PAGE_W / n_cols

                # Banda de rotación con fechas
                titulo_rot = f"Rotación {rotacion}:   {fecha_ini}  →  {fecha_fin}"
                rot_tbl = Table([[Paragraph(titulo_rot, rot_s)]], colWidths=[PAGE_W])
                rot_tbl.setStyle(TableStyle([
                    ('BACKGROUND',    (0,0), (-1,-1), AZUL_ROTACION),
                    ('TOPPADDING',    (0,0), (-1,-1), 5),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                    ('LEFTPADDING',   (0,0), (-1,-1), 8),
                ]))
                elements.append(rot_tbl)

                # Tabla: escenarios en columnas, estudiantes en filas
                header_row  = [Paragraph(e, esc_s) for e in escenarios_orden]
                docente_row = [Paragraph(esc_data[e]["docente"], doc_s) for e in escenarios_orden]

                max_est = max(len(esc_data[e]["estudiantes"]) for e in escenarios_orden)
                student_rows = []
                for i in range(max_est):
                    student_rows.append([
                        Paragraph(
                            esc_data[e]["estudiantes"][i] if i < len(esc_data[e]["estudiantes"]) else "",
                            est_s
                        )
                        for e in escenarios_orden
                    ])

                data_tbl = [header_row, docente_row] + student_rows
                ts = [
                    ('BACKGROUND', (0,0), (-1,0), AZUL_HEADER),
                    ('TEXTCOLOR',  (0,0), (-1,0), BLANCO),
                    ('BACKGROUND', (0,1), (-1,1), VERDE_DOCENTE),
                    ('GRID',    (0,0), (-1,-1), 0.5, colors.HexColor('#BFBFBF')),
                    ('ALIGN',   (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN',  (0,0), (-1,-1), 'MIDDLE'),
                    ('TOPPADDING',    (0,0), (-1,-1), 3),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ]
                for ri in range(2, len(data_tbl)):
                    ts.append(('BACKGROUND', (0,ri), (-1,ri),
                               GRIS_ALTERNO if ri % 2 == 0 else BLANCO))

                tbl = Table(data_tbl, colWidths=[col_w]*n_cols)
                tbl.setStyle(TableStyle(ts))
                elements.append(tbl)
                elements.append(Spacer(1, 10))

        doc.build(elements)
        buffer.seek(0)

        return send_file(buffer, mimetype='application/pdf',
                         as_attachment=True,
                         download_name='Programacion_Practicas_2026-1.pdf')

    

    except Exception as e:
        flash(f'Error generando PDF: {str(e)}', 'danger')
        return redirect(url_for('index'))


@app.route('/auto_assignment', methods=['GET', 'POST'])
def auto_assignment():
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        try:
            estudiantes = request.form.getlist('estudiantes')  # lista de IDs
            docentes = request.form.getlist('docentes')        # lista de IDs
            escenarios = request.form.getlist('escenarios')    # lista de IDs
            horario = request.form['horario'].strip()
            fecha_inicio = request.form['fecha_inicio']
            fecha_fin = request.form['fecha_fin']

            total_rotaciones = len(escenarios)

            # Rotación 1: asignar estudiantes a escenarios en orden
            asignaciones_r1 = []
            docentes_por_escenario = {}  # mapa escenario -> docente

            for i, est in enumerate(estudiantes):
                esc = escenarios[i % len(escenarios)]
                doc = docentes[i % len(docentes)]
                asignaciones_r1.append((est, esc))
                docentes_por_escenario[esc] = doc  # docente fijo en su escenario

                cur.execute('''
                    INSERT INTO asignaciones (estudiante_id, docente_id, escenario_id, rotacion, horario, fecha_inicio, fecha_fin)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                ''', (est, doc, esc, 1, horario, fecha_inicio, fecha_fin))

            # Generar rotaciones automáticas
            from datetime import datetime, timedelta
            fecha_ini = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

            # Duración en días de cada rotación
            duracion_rotacion = (fecha_fin_dt - fecha_ini).days + 1

            for r in range(2, total_rotaciones + 1):
                # inicio = día siguiente al fin de la rotación anterior
                fecha_ini_r = fecha_ini + timedelta(days=duracion_rotacion * (r-1))
                # fin = inicio + duración - 1
                fecha_fin_r = fecha_ini_r + timedelta(days=duracion_rotacion - 1)

                for est, esc in asignaciones_r1:
                    nuevo_esc = escenarios[(escenarios.index(esc) + (r-1)) % total_rotaciones]
                    docente_fijo = docentes_por_escenario[nuevo_esc]  # docente permanece en su escenario

                    cur.execute('''
                        INSERT INTO asignaciones (estudiante_id, docente_id, escenario_id, rotacion, horario, fecha_inicio, fecha_fin)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''', (est, docente_fijo, nuevo_esc, r, horario, fecha_ini_r.date(), fecha_fin_r.date()))

            conn.commit()
            flash('✅ Rotaciones generadas automáticamente con fechas consecutivas y docentes fijos en sus escenarios', 'success')
            return redirect(url_for('asignaciones_list'))

        except Exception as e:
            flash(f'❌ Error: {str(e)}', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()

    # GET - cargar listas
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, cedula FROM estudiantes ORDER BY nombre")
    estudiantes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM docentes WHERE estado = 'Activo' ORDER BY nombre")
    docentes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM escenarios WHERE estado = 'Activo' ORDER BY nombre")
    escenarios = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('auto_assignment.html',
                           estudiantes=estudiantes,
                           docentes=docentes,
                           escenarios=escenarios)


@app.route('/delete_all_assignments')
def delete_all_assignments():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Eliminar todas las asignaciones
        cur.execute("DELETE FROM asignaciones")
        conn.commit()
        
        flash('🗑️ Todas las asignaciones fueron eliminadas correctamente', 'danger')
        return redirect(url_for('asignaciones_list'))
        
    except Exception as e:
        flash(f'❌ Error al eliminar todas las asignaciones: {str(e)}', 'danger')
        if conn:
            conn.rollback()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

@app.route('/search_assignments')
def search_assignments():
    query = request.args.get('q', '').strip()
    conn = get_db_connection()
    cur = conn.cursor()

    # Buscar por nombre de estudiante o número de documento
    cur.execute('''
        SELECT a.id, e.nombre AS estudiante, e.cedula, d.nombre AS docente, 
               s.nombre AS escenario, a.rotacion, a.horario, 
               a.fecha_inicio, a.fecha_fin
        FROM asignaciones a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN docentes d ON a.docente_id = d.id
        JOIN escenarios s ON a.escenario_id = s.id
        WHERE e.nombre ILIKE %s OR e.cedula ILIKE %s
        ORDER BY a.fecha_inicio
    ''', (f'%{query}%', f'%{query}%'))

    resultados = cur.fetchall()
    cur.close()
    conn.close()

    if not resultados:
        flash('⚠️ No se encontraron asignaciones con ese criterio', 'warning')

    return render_template('asignaciones.html', asignaciones=resultados)


        
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
